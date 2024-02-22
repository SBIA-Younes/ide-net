import openpyxl
import json
from datetime import datetime

# Charger le fichier Excel existant
wb = openpyxl.load_workbook('./data/sort all-xlsx.xlsx')

# Sélectionner la feuille de calcul avec laquelle vous souhaitez travailler
sheet = wb['Collecte']  # Ou vous pouvez sélectionner une feuille spécifique par son nom : wb['Nom_de_la_feuille']

# Créer une liste pour stocker les données

data = []

# Obtenir le nombre de colonnes dans la feuille de calcul
num_columns = sheet.max_column

# Boucler à travers chaque ligne et colonne pour récupérer les données
for row in sheet.iter_rows(values_only=True):
    # Créer un dictionnaire pour stocker les données de chaque ligne
    row_data = {}
    for col_idx in range(num_columns):
        # Utiliser l'en-tête de colonne comme clé dans le dictionnaire
        col_header = sheet.cell(row=1, column=col_idx + 1).value
        cell_value = row[col_idx]
        # Vérifier si la valeur est de type datetime
        if isinstance(cell_value, datetime):
            # Formater la valeur datetime en chaîne de caractères
            cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(cell_value, float) and cell_value.is_integer():
            # Si la valeur est un nombre flottant et qu'elle est entière, la convertir en entier
            cell_value = int(cell_value)
        else:
            # Convertir les autres valeurs en chaînes de caractères
            cell_value = str(cell_value)
            # Supprimer le ".0" si présent
            cell_value = cell_value.rstrip('.0')
        row_data[col_header] = cell_value
    # Ajouter le dictionnaire à la liste des données
    data.append(row_data)

# Fermer le fichier Excel après avoir fini de le lire
wb.close()

# Écrire les données dans un fichier JSON
with open('./data/donneesCollecte.json', 'w', encoding='utf-8') as json_file:
    json.dump(data, json_file, indent=4, ensure_ascii=False)

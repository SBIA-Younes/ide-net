import openpyxl
import json

wb = openpyxl.load_workbook('./data/FLOTTE FBA Au 31-12-2023.xlsx')

sheet = wb['CORSO']

# Créer une liste pour stocker les données
data = []

for row in sheet.iter_rows(values_only=True):
    # Créer un dictionnaire pour stocker les données de chaque ligne
    row_data = {
       "N°": row[0],
        "N° INTERNE TRACTEUR": row[1],
        "TAG": row[2],
        "NOM CHAUFFEUR": row[3],
        "TEL FBA": row[4],
        "N° PERMIS CHAUFFEUR": row[5],
        "Permis délivré le": row[6],
        "VALIDITE PERMIS CHAUFFEUR": row[7],
        "Expération PC dans": row[8],
        "MATRICULE TRACTEUR": row[9],
        "age": row[10],
        "Date début CT": row[11],
        "VALIDITE CONTRÔLE TECHNIQUE": row[12],
        "Jours Restants CT": row[13],
        "VALIDITE PERMIS A CIRCULER": row[14],
        "Jours Restants PAC": row[15],
        "NUMERO DE CHASSIS TRACTEUR": row[16],
        "TYPE TRACTEUR": row[17],
        "NOMBRE SANGLE": row[18],
        "N° Interne Extincteur": row[19],
        "VALIDITE EXTINCTEUR": row[20],
        "Experation Extincteur Dans": row[21],
        "TRIANGLE": row[22],
        "BOITE A PHARMACIE": row[23],
        "CALE DE ROUE": row[24],
        "CLE CABINE":  row[25],
        "Clé de Roue": row[26],
        "CRIQUE":  row[27],
        "Clé 27":  row[28],
        "NUMERO INTERNE REMORQUE": row[29],
        "Camion de la remorque":  row[30],
        "TYPE REMORQUE":row[31],
        "MATRICULE REMORQUE": row[32],
        "null": row[33],
        "N° CHASSIS REMORQUE": row[34],
        "Pneus Neuf": row[35],
        "NOMBRE PLANCHE": row[36],
        "Roue de Secours": row[37],
        "NOMBRE EQUERRE": row[38],
        "BACHE PLATEAU": row[39],
        "DATE DE MISE A JOUR": row[40]
    }
    # Ajouter le dictionnaire à la liste des données
    data.append(row_data)

# Fermer le fichier Excel après avoir fini de le lire
wb.close()

# Écrire les données dans un fichier JSON
with open('./data/tout.json', 'w', encoding='utf-8') as json_file:
    json.dump(data, json_file, indent=4, ensure_ascii=False)

# Obtenir le nombre de colonnes dans la feuille de calcul
# num_columns = sheet.max_column

# # Boucler à travers chaque ligne et colonne pour récupérer les données
# for row in sheet.iter_rows(values_only=True):
#     # Créer un dictionnaire pour stocker les données de chaque ligne
#     row_data = {}
#     for col_idx in range(num_columns):
#         # Utiliser l'en-tête de colonne comme clé dans le dictionnaire
#         col_header = sheet.cell(row=1, column=col_idx + 1).value
#         row_data[col_header] = row[col_idx]
#     # Ajouter le dictionnaire à la liste des données
#     data.append(row_data)

# # Fermer le fichier Excel après avoir fini de le lire
# wb.close()

# # Écrire les données dans un fichier JSON
# with open('donnees_2.json', 'w', encoding='utf-8') as json_file:
#     json.dump(data, json_file, indent=4, ensure_ascii=False)